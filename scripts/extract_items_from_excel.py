#!/usr/bin/env python3
"""
Script para extrair itens de avaliação da planilha GERADOR DE OIs
e gerar JSON para importação no sistema
"""

import openpyxl
import json
import sys

def extract_items_from_excel(excel_file):
    """Extrai itens da aba DIST da planilha Excel"""
    
    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as e:
        print(f"Erro ao carregar arquivo: {e}", file=sys.stderr)
        return None
    
    if 'DIST' not in wb.sheetnames:
        print("Erro: Aba 'DIST' não encontrada na planilha", file=sys.stderr)
        return None
    
    ws = wb['DIST']
    items = []
    
    # Começar a ler a partir da linha 4 (linha 1-3 são cabeçalhos)
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Verificar se a linha tem dados
        if not any(cell is not None for cell in row):
            continue
        
        # Estrutura esperada:
        # Coluna A (índice 0): REFERÊNCIA
        # Coluna B (índice 1): ANEXO-C (categoria)
        # Coluna C (índice 2): OI (descrição do item)
        # Colunas D+ (índices 3+): Distribuição (x indica que o item é avaliado naquele estágio)
        
        anexo_c = row[1]  # Coluna B
        oi = row[2]       # Coluna C
        referencia = row[0]  # Coluna A
        
        if not anexo_c or not oi:
            continue
        
        # Extrair distribuição (colunas D até o final)
        distribuicao = []
        for col_idx in range(3, len(row)):
            if row[col_idx] == 'x':
                # Adicionar o número do estágio (1-indexed)
                distribuicao.append(col_idx - 2)
        
        item = {
            "anexoC": str(anexo_c).strip(),
            "oi": str(oi).strip(),
            "referencia": str(referencia).strip() if referencia else None,
            "distribuicao": distribuicao,
        }
        
        items.append(item)
    
    return items

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 extract_items_from_excel.py <arquivo_excel>", file=sys.stderr)
        sys.exit(1)
    
    excel_file = sys.argv[1]
    items = extract_items_from_excel(excel_file)
    
    if items is None:
        sys.exit(1)
    
    # Imprimir JSON
    print(json.dumps({"items": items}, indent=2, ensure_ascii=False))

if __name__ == "__main__":
    main()
